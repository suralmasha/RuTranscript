import warnings
from pathlib import Path

import epitran
import spacy
from nltk.stem.snowball import SnowballStemmer
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from tps import download, find
from tps import modules as md

from .tools import (
    SyntaxTree,
    allophones,
    apply_differences,
    assimilative_palatalization,
    epi_starterpack,
    find_clitics,
    first_jot,
    fix_jotised,
    get_punctuation_dict,
    labia_velar,
    long_consonants,
    long_ge,
    merge_phrasal_words,
    nasal_m_n,
    put_stresses,
    remove_extra_stresses,
    replace_stress_before,
    shch,
    silent_r,
    stunning,
    text_norm_tok,
    voiced_ts,
    vowels,
)

snowball = SnowballStemmer('russian')
nlp = spacy.load('ru_core_news_sm', disable=['tagger', 'morphologizer', 'attribute_ruler'])

ROOT_DIR: Path = Path(__file__).resolve().parent
wb: Workbook = load_workbook(ROOT_DIR / 'data' / 'irregular_exceptions.xlsx')

sheet = wb.active
irregular_exceptions = {sheet[f'A{i}'].value: sheet[f'B{i}'].value for i in range(2, sheet.max_row + 1)}
irregular_exceptions_stems = {snowball.stem(ex): pron for ex, pron in irregular_exceptions.items()}

epi = epitran.Epitran('rus-Cyrl')
second_silent = ['стн', 'стл', 'здн', 'рдн', 'нтск', 'ндск', 'лвств']
first_silent = ['лнц', 'дц', 'вств']
hissing_rd = {'сш': 'шш', 'зш': 'шш', 'сж': 'жж', 'сч': 'щ'}
non_ipa_symbols = {'t͡ɕʲ': 't͡ɕ', 'ʂʲː': 'ʂ', 'ɕːʲ': 'ɕː', 'ʒ': 'ʐ', 'd͡ʐ': 'd͡ʒ'}

try:
    yo_dict = find('yo.dict', raise_exception=True)
except FileNotFoundError:
    yo_dict = download('yo.dict')

try:
    e_dict = find('e.dict', raise_exception=True)
except FileNotFoundError:
    e_dict = download('e.dict')

e_replacer = md.Replacer([e_dict, 'plane'])
yo_replacer = md.Replacer([yo_dict, 'plane'])
syntax_tree = SyntaxTree()


class RuTranscript:
    """
    Russian transcription processor.

    This class provides methods to convert Russian text into its phonetic transcription.
    """

    def __init__(
        self,
        text: str,
        stressed_text: str | None = None,
        stress_place: str = 'after',
        replacement_dict: dict | None = None,
        stress_accuracy_threshold: float = 0.86,
    ) -> None:
        """
        Make a phonetic transcription in russian using IPA.

        :param text: A text to transcribe.
        :param stressed_text: The same (!) text with stresses.
            You may define stresses both for one word and for all words in the text.
            To do this, put a stress symbol (preferably '+') before or after the stressed vowel.
        :param stress_place: 'after' - if the stress symbol is after the stressed vowel,
            'before' - if the stress symbol is before the stressed vowel.
        :param replacement_dict: Custom dictionary for replacing words (for example, {'tts': 'синтез речи'}).
        :param stress_accuracy_threshold: A threshold for the accuracy of stress placement for StressRNN.
        """
        text, stressed_text = self._get_text_and_stressed_text(text, stressed_text, replacement_dict)
        self._pause_dict = get_punctuation_dict(text)
        self._tokens = text_norm_tok(text)
        self._sections_len = len(self._tokens)
        self._stressed_tokens = text_norm_tok(stressed_text)

        self._stress_accuracy_threshold = stress_accuracy_threshold
        self._stress_place = stress_place

        self._phrasal_words_indexes = []
        self._letters_list = []
        self._phonemes_list = []
        self._allophones_list = [[]] * self._sections_len
        self._transliterated_tokens = [[]] * self._sections_len
        self._phrasal_words = [[]] * self._sections_len
        self._stressed_text = [[]] * self._sections_len

    @staticmethod
    def _get_text_and_stressed_text(
        text: str, stressed_text: str | None, replacement_dict: dict | None
    ) -> tuple[str, str]:
        """
        Prepare and normalizes text and its stressed version.

        Replaces newlines with spaces, converts text to lowercase,
        and replaces '-' with '—'. Optionally applies a replacement dictionary.

        param text: Original text.
        param stressed_text: Text with stress marks, or None to use `text`.
        param replacement_dict: Optional dictionary for custom replacements.
        return: Tuple of normalized text and stressed text.
        """
        text = ' '.join(['—' if word == '-' else word for word in text.replace('\n', ' ').lower().split()])
        stressed_text = (
            ' '.join(['—' if word == '-' else word for word in stressed_text.replace('\n', ' ').lower().split()])
            if stressed_text is not None
            else text
        )

        if replacement_dict is not None:
            user_replacer = md.Replacer([replacement_dict, 'plane'])
            text = user_replacer(text)
            stressed_text = user_replacer(stressed_text)

        return text, stressed_text

    def _remove_dashes(self, section_num: int) -> None:
        """
        Remove hyphens from tokens in the specified section.

        param section_num: Index of the section to process.
        """
        section = self._tokens[section_num]
        a_section = self._stressed_tokens[section_num]
        self._tokens[section_num] = [token.replace('-', '') for token in section]
        self._stressed_tokens[section_num] = [
            token.replace('-', '') if token.count('+') == 1 else remove_extra_stresses(token).replace('-', '')
            for token in a_section
        ]

    def _tps(self, section_num: int) -> None:
        """
        Replace 'е - э' and 'е - ё' in tokens of the specified section and update stressed tokens.

        param section_num: Index of the section to process.
        """
        default_section = self._tokens[section_num]
        self._tokens[section_num] = [e_replacer(token.replace('+', '')) for token in self._tokens[section_num]]
        self._tokens[section_num] = [yo_replacer(token.replace('+', '')) for token in self._tokens[section_num]]

        if self._tokens[section_num] != [token.replace('+', '') for token in default_section]:
            self._stressed_tokens[section_num] = [
                apply_differences([default_section[i], self._tokens[section_num][i]])
                for i in range(len(default_section))
            ]

    @staticmethod
    def _join_phonemes(transliterated_tokens: list[str], limit: int = 10000) -> list[str]:
        """
        Join transliterated tokens into a list of phonemes.

        param transliterated_tokens: List of tokens to convert into phonemes.
        param limit: Maximum allowed iterations to prevent endless loops.
        return: List of phonemes.
        """
        section_phonemes_list = []
        joined_tokens = '_'.join(transliterated_tokens)
        joined_tokens = joined_tokens.replace('‑', '-')
        i = 0
        counter = 0
        default_len = len(joined_tokens)
        while i < default_len:
            if joined_tokens[i] not in ['+', '-']:
                n = 4
                if i != default_len - 1:
                    while (joined_tokens[i : i + n] not in [*epi_starterpack, '_', '|', '||', 'γ', 'ʐ']) and (n > 0):
                        counter += 1
                        if counter > limit:
                            raise IndexError('Endless loop')  # noqa: TRY003
                        n -= 1
                    section_phonemes_list.append(joined_tokens[i : i + n])
                elif joined_tokens[i] in [*epi_starterpack, '||', 'γ']:
                    section_phonemes_list.append(joined_tokens[i])
                i += n
            else:
                section_phonemes_list.append(joined_tokens[i])
                i += 1

        section_phonemes_list = [x for x in section_phonemes_list if x not in ['', 'ʲ']]

        n = 0
        for allophone_index in range(len(section_phonemes_list) - 1):
            allophone = section_phonemes_list[allophone_index + n]
            next_allophone = section_phonemes_list[allophone_index + n + 1]
            if (allophone == 't͡s' and next_allophone == 's') or (allophone == 'd͡ʒ' and next_allophone == 'ʐ'):
                del section_phonemes_list[allophone_index + n + 1]
                n -= 1

        # print(section_phonemes_list)
        return section_phonemes_list

    @staticmethod
    def add_prestressed_syllable_sign(section: list[str]) -> list[str]:
        """
        Insert a prestressed syllable mark ('-') before a stressed vowel.

        param section: List of phonemes with stress markers ('+').
        return: List of phonemes with prestressed syllable marks added.
        """
        section_result = section[:]
        n = 0
        for symb_i, symb in enumerate(section):
            if symb == '+':
                preavi = [
                    phon_i
                    for phon_i, phon in enumerate(section[: symb_i - 1])
                    if allophones[phon]['phon'] == 'V' and '_' not in section[phon_i + n : symb_i]
                ]
                if preavi:
                    section_result.insert(preavi[-1] + n + 1, '-')
                    n += 1

        return section_result

    def _lpt_1(self, section_num: int) -> None:
        """
        Letter-to-phoneme transformation by B.M. Lobanov. Part 1 - Irregular exceptions.

        param section_num: Index of the section to process.
        return: None. Updates `_tokens` and `_stressed_tokens` in place.
        """
        for i, token in enumerate(self._tokens[section_num]):
            stem = snowball.stem(token)
            if stem in irregular_exceptions_stems:
                try:
                    new_token = irregular_exceptions[token]
                except KeyError:
                    ending = token[len(stem) :]
                    dif = -(len(token) - len(stem))
                    new_token = irregular_exceptions_stems[stem][:dif] + ending

                self._tokens[section_num][i] = new_token
                accent_index = self._stressed_tokens[section_num][i].index('+')
                self._stressed_tokens[section_num][i] = new_token[:accent_index] + '+' + new_token[accent_index:]

    def _lpt_2(self, section_num: int) -> None:  # noqa: PLR0912
        """
        Letter-to-phoneme transformation by B.M. Lobanov. Part 2 - Regular exceptions.

        param section_num: Index of the section to process.
        return: None. Updates `_tokens` and `_stressed_tokens` in place.
        """
        for i, token in enumerate(self._stressed_tokens[section_num]):
            # adjective endings 'ого его'
            if token != 'ого+' and (  # noqa: S105
                token.replace('+', '').startswith('какого')
                or token.replace('+', '').endswith('ого')
                or token.replace('+', '').endswith('его')
            ):
                accent_index = token.index('+')
                token = token.replace('+', '').replace('ого', 'ово').replace('его', 'ево')  # noqa: PLW2901
                self._stressed_tokens[section_num][i] = token[:accent_index] + '+' + token[accent_index:]

            # 'что' --> 'што'
            if 'что' in self._stressed_tokens[section_num][i]:
                self._stressed_tokens[section_num][i] = token.replace('что', 'што')

            # verb endings 'тся ться'
            if token not in {'заботься', 'отметься'}:
                if token[-3:] == 'тся':
                    self._stressed_tokens[section_num][i] = token[:-3] + 'ца'
                elif token[-4:] == 'ться':
                    self._stressed_tokens[section_num][i] = token[:-4] + 'ца'

            # noun endings 'ия ие ию'
            if (token[-2:] in {'ия', 'ие', 'ию'}) and (token[-3] not in {'ц', 'щ'}):
                if token[-3] not in {'ж', 'ш'}:
                    self._stressed_tokens[section_num][i] = token[:-2] + 'ь' + token[-1]
                else:
                    self._stressed_tokens[section_num][i] = token[:-2] + 'й' + token[-1]

            # unpronounceable consonants
            for sub in first_silent + second_silent:
                if sub in token:
                    new_sub = sub.translate(str.maketrans('', '', 'ьъ'))
                    self._stressed_tokens[section_num][i] = token.translate(str.maketrans(sub, new_sub))

            # combinations with hissing consonants
            stem = snowball.stem(token)
            if ('зч' in token or 'тч' in token or 'дч' in token) and (stem[-3:] == 'чик' or stem[-3:] == 'чиц'):
                self._stressed_tokens[section_num][i] = token.replace('зч', 'щ').replace('тч', 'ч').replace('дч', 'ч')
            for key, value in hissing_rd.items():
                if key in token:
                    self._stressed_tokens[section_num][i] = token.replace(key, value)

    def _lpt_3(self, section_num: int) -> None:
        """
        Letter-to-phoneme transformation by B.M. Lobanov. Part 3 - Transliteration.

        param section_num: Index of the section to process.
        return: None. Updates `_transliterated_tokens` in place.
        """
        self._transliterated_tokens[section_num] = [
            epi.transliterate(token).replace('6', '').replace('4', '') for token in self._stressed_tokens[section_num]
        ]
        for i, token in enumerate(self._transliterated_tokens[section_num]):
            for key, value in non_ipa_symbols.items():
                if key in token:
                    token = token.replace(key, value)  # noqa: PLW2901
                    self._transliterated_tokens[section_num][i] = token

    def _lpt_4(self, section_num: int) -> None:
        """
        Letter-to-phoneme transformation by B.M. Lobanov. Part 4 - Common Rules.

        param section_num: Index of the section to process.
        return: None. Updates `_transliterated_tokens` in place.
        """
        # fricative g
        for i, token in enumerate(self._transliterated_tokens[section_num]):
            try:
                next_token = self._transliterated_tokens[section_num][i + 1]
            except IndexError:
                next_token = ' '  # noqa: S105

            token_let = self._tokens[section_num][i]
            nlp_token = nlp(token_let)[0]
            lemma = nlp_token.lemma_

            if lemma in {'ага', 'ого', 'угу', 'господь', 'господи', 'бог'}:
                self._transliterated_tokens[section_num][i] = token.replace('ɡ', 'γ', 1)
            elif token_let in {'ах', 'эх', 'ох', 'ух'}:
                next_token_allophone = allophones.get(next_token[0], {})
                if next_token_allophone.get('voice', '') == 'voiced':
                    self._transliterated_tokens[section_num][i] = token.replace('x', 'γ', 1)

        # ---- Join phonemes ----
        joined_phonemes = self._join_phonemes(self._transliterated_tokens[section_num], limit=10000)
        self._phonemes_list.append(joined_phonemes)

        # ---- Join letters ----
        joined_letters = list('_'.join(self._stressed_tokens[section_num]))
        self._letters_list.append(joined_letters)

        # ---- Continue LPC-4. Common rules ----
        self._phonemes_list[section_num] = fix_jotised(
            self._phonemes_list[section_num], self._letters_list[section_num]
        )
        self._phonemes_list[section_num] = shch(self._phonemes_list[section_num])
        self._phonemes_list[section_num] = long_ge(self._phonemes_list[section_num])
        self._phonemes_list[section_num] = assimilative_palatalization(
            self._tokens[section_num], self._phonemes_list[section_num]
        )
        self._phonemes_list[section_num] = long_consonants(self._phonemes_list[section_num])
        self._phonemes_list[section_num] = stunning(self._phonemes_list[section_num])

    def transcribe(self) -> None:
        """
        Perform full transcription pipeline on all sections.

        return: None. Updates internal token, stressed token, phoneme, and allophone lists in place.
        """
        for section_num in range(self._sections_len):
            self._tps(section_num)
            # ---- Accenting ----
            self._stressed_tokens[section_num] = put_stresses(
                tokens_list=self._stressed_tokens[section_num],
                stress_place=self._stress_place,
                stress_accuracy_threshold=self._stress_accuracy_threshold,
            )
            self._stressed_text[section_num] = self._stressed_tokens[section_num]
            # ---- Removing dashes ----
            self._remove_dashes(section_num)
            # ---- Phrasal words extraction ----
            dep = syntax_tree.make_dependency_tree(' '.join(self._tokens[section_num]))
            self._phrasal_words_indexes.append(find_clitics(dep, self._tokens[section_num]))
            # ---- Letter-phoneme transformation ----
            self._lpt_1(section_num)
            self._lpt_2(section_num)
            self._lpt_3(section_num)
            self._lpt_4(section_num)
            # ---- Allophones - consonants ----
            self._allophones_list[section_num] = first_jot(self._phonemes_list[section_num])
            self._allophones_list[section_num] = nasal_m_n(self._allophones_list[section_num])
            self._allophones_list[section_num] = silent_r(self._allophones_list[section_num])
            self._allophones_list[section_num] = voiced_ts(self._allophones_list[section_num])
            # ---- Extract phrasal words ----
            self._phrasal_words[section_num] = merge_phrasal_words(
                self._allophones_list[section_num], self._phrasal_words_indexes[section_num]
            )
            #  ---- Allophones - vowels ----
            self._phrasal_words[section_num] = self.add_prestressed_syllable_sign(self._phrasal_words[section_num])
            self._allophones_list[section_num] = vowels(self._phrasal_words[section_num])
            self._allophones_list[section_num] = labia_velar(self._allophones_list[section_num])

    def _insert_pauses(self, sounds_list: list) -> None:
        """
        Insert pauses into the sounds list according to the pause dictionary.

        param sounds_list: List of phonemes or allophones where pauses will be inserted.
        return: None. Modifies sounds_list in place.
        """
        for i, key in enumerate(self._pause_dict):
            sounds_list.insert(i + key, self._pause_dict[key])

    @staticmethod
    def _get_escape_symbols(save_stresses: bool = False, save_spaces: bool = False) -> list[str]:
        """
        Get symbols to escape from removal or processing.

        param save_stresses: Keep '+' symbols if True.
        param save_spaces: Keep '_' symbols if True.
        return: List of symbols to escape.
        """
        escape_symbols = ['+', '-', '_']
        if save_stresses:
            escape_symbols.remove('+')
        if save_spaces:
            escape_symbols.remove('_')

        return escape_symbols

    @staticmethod
    def _join_sounds(escape_symbols: list, sounds_list: list) -> str:
        """
        Join list of sounds into a single string, excluding escape symbols.

        param escape_symbols: List of symbols to skip during joining.
        param sounds_list: List of phoneme/allophone sequences or pause markers.
        return: Joined string of sounds.
        """
        return ' '.join(
            [
                ' '.join([x for x in section if x not in escape_symbols]) if section != '||' else section
                for section in sounds_list
            ]
        )

    def get_allophones(
        self,
        stress_place: str | None = None,
        save_stresses: bool = False,
        save_spaces: bool = False,
        save_pauses: bool = False,
        stress_symbol: str = '+',
    ) -> list[str]:
        """
        Return a list of allophones.

        :param stress_place: 'after' - to place the stress symbol after the stressed vowel,
            'before' - to place the stress symbol before the stressed vowel.
        :param stress_symbol: A symbol that you want to indicate the stress.
            Be careful not to use letters and signs from the following list
            ['.', '_', '-', 'ʲ', 'ᶣ', 'ʷ', 'ˠ', 'ː', '͡']!
        :param save_spaces: Will replace spaces with '_'.
        :param save_stresses: Will replace stresses with the stress_symbol.
        :param save_pauses: Will replace punctuation with '||' for long pauses ('.', '?', '!', '…')
            and '|' for short pauses (other symbols).
        :return: List of allophones.
        """
        if stress_symbol in ['.', '_', '-', 'ʲ', 'ᶣ', 'ʷ', 'ˠ', 'ː', '͡']:
            warnings.warn(
                'The stress symbol intersects with the IPA transcription signs '
                'or the internal sighs of the framework.\nIt may cause an unpredictable behaviour.\n'
                "Better don't use signs from the following list "
                "['.', '_', '-', 'ʲ', 'ᶣ', 'ʷ', 'ˠ', 'ː', '͡']!",
                stacklevel=2,
            )

        if save_pauses:
            self._insert_pauses(self._allophones_list)

        escape_symbols = self._get_escape_symbols(save_stresses=save_stresses, save_spaces=save_spaces)
        res = self._join_sounds(escape_symbols, self._allophones_list).split()

        if stress_place is None:
            stress_place = self._stress_place
        if stress_place == 'before':
            res = replace_stress_before(res)

        if (stress_symbol != '+') and ('+' not in escape_symbols):
            res = [x.replace('+', stress_symbol) for x in res]

        return res

    def get_phonemes(
        self,
        stress_place: str | None = None,
        save_stresses: bool = False,
        save_spaces: bool = False,
        save_pauses: bool = False,
        stress_symbol: str = '+',
    ) -> list[str]:
        """
        Return a list of phonemes.

        :param stress_place: 'after' - to place the stress symbol after the stressed vowel,
            'before' - to place the stress symbol before the stressed vowel.
        :param stress_symbol: A symbol that you want to indicate the stress.
            Be careful not to use letters and signs from the following list
            ['.', '_', '-', 'ʲ', 'ᶣ', 'ʷ', 'ˠ', 'ː', '͡']!
        :param save_spaces: Will replace spaces with '_'.
        :param save_stresses: Will replace stresses with the stress_symbol.
        :param save_pauses: Will replace punctuation with '||' for long pauses ('.', '?', '!', '…')
            and '|' for short pauses (other symbols).
        :return: List of phonemes.
        """
        if stress_symbol in ['.', '_', '-', 'ʲ', 'ᶣ', 'ʷ', 'ˠ', 'ː', '͡']:
            warnings.warn(
                'The stress symbol intersects with the IPA transcription signs '
                'or the internal sighs of the framework.\nIt may cause an unpredictable behaviour.\n'
                "Better don't use signs from the following list "
                "['.', '_', '-', 'ʲ', 'ᶣ', 'ʷ', 'ˠ', 'ː', '͡']!",
                stacklevel=2,
            )

        if save_pauses:
            self._insert_pauses(self._phonemes_list)

        escape_symbols = self._get_escape_symbols(save_stresses=save_stresses, save_spaces=save_spaces)
        res = self._join_sounds(escape_symbols, self._phonemes_list).split()

        if stress_place is None:
            stress_place = self._stress_place
        if stress_place == 'before':
            res = replace_stress_before(res)

        if (stress_symbol != '+') and ('+' not in escape_symbols):
            res = [x.replace('+', stress_symbol) for x in res]

        return res

    def get_stressed_text(self, stress_place: str | None = None, stress_symbol: str = '+') -> str:
        """
        Return text with stress markers.

        :param stress_place: 'after' - to place the stress symbol after the stressed vowel,
            'before' - to place the stress symbol before the stressed vowel.
        :param stress_symbol: A symbol that you want to indicate the stress.
            Be careful not to use signs from the following list ['.', '_', '-', 'ʲ', 'ᶣ', 'ʷ', 'ˠ', 'ː', '͡']!
        :return: A text string with stresses.
        """
        if stress_symbol in ['.', '_', '-', 'ʲ', 'ᶣ', 'ʷ', 'ˠ', 'ː', '͡']:
            warnings.warn(
                'The stress symbol intersects with the IPA transcription signs '
                'or the internal sighs of the framework.\nIt may cause an unpredictable behaviour.\n'
                "Better don't use signs from the following list "
                "['.', '_', '-', 'ʲ', 'ᶣ', 'ʷ', 'ˠ', 'ː', '͡']!",
                stacklevel=2,
            )

        if stress_place is None:
            stress_place = self._stress_place
        if stress_place == 'before':
            res = ' '.join([''.join(replace_stress_before(' '.join(section))) for section in self._stressed_text])
        else:
            res = ' '.join([' '.join(section) for section in self._stressed_text])

        if stress_symbol != '+':
            res = res.replace('+', stress_symbol)

        return res
