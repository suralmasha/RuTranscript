from os.path import join, dirname, abspath

import spacy
import epitran
from openpyxl import load_workbook
from nltk.stem.snowball import SnowballStemmer
from tps import find, download
from tps import modules as md

from ru_transcript.src.utils.main_tools import get_punctuation_dict, text_norm_tok, SyntaxTree, find_clitics, \
    extract_phrasal_words, apply_differences, place_stress, replace_stress, remove_extra_stresses, replace_stress_before
from ru_transcript.src.utils.sounds import epi_starterpack, allophones
from ru_transcript.src.utils.allophones_tools import nasal_m_n, silent_r, voiced_ts, shch, long_ge, fix_jotised,\
    assimilative_palatalization, long_consonants, vowels, labia_velar, stunning

snowball = SnowballStemmer('russian')
nlp = spacy.load('ru_core_news_sm', disable=["tagger", "morphologizer", "attribute_ruler"])

ROOT_DIR = dirname(abspath(__file__))
wb = load_workbook(join(ROOT_DIR, 'data/irregular_exceptions.xlsx'))
sheet = wb.active
irregular_exceptions = {sheet[f'A{i}'].value: sheet[f'B{i}'].value for i in range(2, sheet.max_row + 1)}
irregular_exceptions_stems = {snowball.stem(ex): pron for ex, pron in irregular_exceptions.items()}

epi = epitran.Epitran('rus-Cyrl')
second_silent = 'стн стл здн рдн нтск ндск лвств'.split()
first_silent = 'лнц дц вств'.split()
hissing_rd = {'сш': 'шш', 'зш': 'шш', 'сж': 'жж', 'сч': 'щ'}
non_ipa_symbols = {'t͡ɕʲ': 't͡ɕ', 'ʂʲː': 'ʂ', 'ɕːʲ': 'ɕː', 'ʒ': 'ʐ', 'd͡ʐ': 'd͡ʒ'}

try:
    yo_dict = find("yo.dict", raise_exception=True)
except FileNotFoundError:
    yo_dict = download("yo.dict")

try:
    e_dict = find("e.dict", raise_exception=True)
except FileNotFoundError:
    e_dict = download("e.dict")

e_replacer = md.Replacer([e_dict, "plane"])
yo_replacer = md.Replacer([yo_dict, "plane"])
syntax_tree = SyntaxTree()


class RuTranscript:
    def __init__(self, text: str, stressed_text: str = None, stress_place: str = 'after', replacement_dict: dict = None,
                 stress_accuracy_threshold: float = 0.86):
        """
        Makes a phonetic transcription in russian using IPA.

        :param text: A text to transcribe.
        :param stressed_text: The same (!) text with stresses.
            You may define stresses both for one word and for all words in the text.
            To do this, put a stress symbol (preferably '+') before or after the stressed vowel.
        :param stress_place: 'after' - if the stress symbol is after the stressed vowel,
            'before' - if the stress symbol is before the stressed vowel.
        :param replacement_dict: Custom dictionary for replacing words (for example, {'tts': 'синтез речи'}).
        :param stress_accuracy_threshold: A threshold for the accuracy of stress placement for StressRNN.
        """
        text, stressed_text = self._get_text_and_stressed_text(text, stressed_text, replacement_dict)
        self._pause_dict = get_punctuation_dict(text)
        self._tokens = text_norm_tok(text)
        self._sections_len = len(self._tokens)
        self._stressed_tokens = text_norm_tok(stressed_text)

        self._stress_accuracy_threshold = stress_accuracy_threshold
        self._stress_place = stress_place

        self._phrasal_words_indexes = []
        self._phonemes = []
        self._letters_list = []
        self._transliterated_tokens = [[]] * self._sections_len
        self._phrasal_words = [[]] * self._sections_len
        self._allophones = [[]] * self._sections_len
        self._stressed_text = [[]] * self._sections_len

    def _get_text_and_stressed_text(self, text, stressed_text, replacement_dict):
        text = ' '.join(['—' if word == '-' else word for word in text.replace('\n', ' ').lower().split()])
        stressed_text = ' '.join(['—' if word == '-' else word
                                  for word in stressed_text.replace('\n', ' ').lower().split()]) \
            if stressed_text is not None else text

        if replacement_dict is not None:
            user_replacer = md.Replacer([replacement_dict, "plane"])
            text = user_replacer(text)
            stressed_text = user_replacer(stressed_text)

        return text, stressed_text

    def _lpt(self, section_num):
        """
        Letter-phoneme transformations by B.M. Lobanov
        """
        ...

    @staticmethod
    def put_stresses(tokens_list: list, stress_place: str = 'after', stress_accuracy_threshold: float = 0.86):
        """
        Puts or replaces stresses.

        :param tokens_list: List of tokens.
        :param stress_place: 'after' - to place the stress symbol after the stressed vowel,
            'before' - to place the stress symbol before the stressed vowel.
        :param stress_accuracy_threshold: A threshold for the accuracy of stress placement for StressRNN.
        :return: List of tokens.
        """
        return [
            replace_stress(token) if ('+' in token) and (stress_place == 'before')  # need to replace
            else place_stress(token, stress_accuracy_threshold) if ('+' not in token)  # use StressRNN
            else token
            for token in tokens_list]

    def _remove_dashes(self, section_num):
        section = self._tokens[section_num]
        a_section = self._stressed_tokens[section_num]
        self._tokens[section_num] = [token.replace('-', '') for token in section]
        self._stressed_tokens[section_num] = [token.replace('-', '') if token.count('+') == 1
                                              else remove_extra_stresses(token).replace('-', '')
                                              for token in a_section]

    def transcribe(self):
        # ---- TPS ----
        for section_num in range(self._sections_len):
            default_section = self._tokens[section_num]
            self._tokens[section_num] = [e_replacer(token.replace('+', '')) for token in self._tokens[section_num]]
            self._tokens[section_num] = [yo_replacer(token.replace('+', '')) for token in self._tokens[section_num]]

            if self._tokens[section_num] != [token.replace('+', '') for token in default_section]:
                self._stressed_tokens[section_num] = [
                    apply_differences([default_section[i], self._tokens[section_num][i]])
                    for i in range(len(default_section))
                ]

            # ---- Accenting ----
            self._stressed_tokens[section_num] = self.put_stresses(
                tokens_list=self._stressed_tokens[section_num],
                stress_place=self._stress_place,
                stress_accuracy_threshold=self._stress_accuracy_threshold)
            self._stressed_text[section_num] = self._stressed_tokens[section_num]

            # ---- Removing dashes ----
            self._remove_dashes(section_num)

            # ---- Phrasal words extraction ----
            dep = syntax_tree.make_dependency_tree(' '.join(self._tokens[section_num]))
            self._phrasal_words_indexes.append(find_clitics(dep, self._tokens[section_num]))

            # ---- LPT-1. Irregular exceptions ----
            for i, token in enumerate(self._tokens[section_num]):
                stem = snowball.stem(token)
                if stem in irregular_exceptions_stems:
                    try:
                        new_token = irregular_exceptions[token]
                    except KeyError:
                        ending = token[len(stem):]
                        dif = - (len(token) - len(stem))
                        new_token = irregular_exceptions_stems[stem][:dif] + ending

                    self._tokens[section_num][i] = new_token
                    accent_index = self._stressed_tokens[section_num][i].index('+')
                    self._stressed_tokens[section_num][i] = new_token[:accent_index] + '+' + new_token[accent_index:]

            # ---- LPC-2. Regular exceptions ----
            for i, token in enumerate(self._stressed_tokens[section_num]):
                # adjective endings 'ого его'
                if token != 'ого+' and (token.replace('+', '').startswith('какого')
                                        or token.replace('+', '').endswith('ого')
                                        or token.replace('+', '').endswith('его')):
                    accent_index = token.index('+')
                    token = token.replace('+', '').replace('ого', 'ово').replace('его', 'ево')
                    self._stressed_tokens[section_num][i] = token[:accent_index] + '+' + token[accent_index:]

                # 'что' --> 'што'
                if 'что' in self._stressed_tokens[section_num][i]:
                    self._stressed_tokens[section_num][i] = token.replace('что', 'што')

                # verb endings 'тся ться'
                if token not in {'заботься', 'отметься'}:
                    if token[-3:] == 'тся':
                        self._stressed_tokens[section_num][i] = token[:-3] + 'ца'
                    elif token[-4:] == 'ться':
                        self._stressed_tokens[section_num][i] = token[:-4] + 'ца'

                # noun endings 'ия ие ию'
                if (token[-2:] in {'ия', 'ие', 'ию'}) and (token[-3] not in {'ц', 'щ'}):
                    if token[-3] not in {'ж', 'ш'}:
                        self._stressed_tokens[section_num][i] = token[:-2] + 'ь' + token[-1]
                    else:
                        self._stressed_tokens[section_num][i] = token[:-2] + 'й' + token[-1]

                # unpronounceable consonants
                for sub in first_silent + second_silent:
                    if sub in token:
                        new_sub = sub.translate(str.maketrans('', '', 'ьъ'))
                        self._stressed_tokens[section_num][i] = token.translate(str.maketrans(sub, new_sub))

                # combinations with hissing consonants
                stem = snowball.stem(token)
                if ('зч' in token or 'тч' in token or 'дч' in token) and (stem[-3:] == 'чик' or stem[-3:] == 'чиц'):
                    self._stressed_tokens[section_num][i] = token.replace('зч', 'щ').replace('тч', 'ч').replace('дч',
                                                                                                                'ч')
                for key, value in hissing_rd.items():
                    if key in token:
                        self._stressed_tokens[section_num][i] = token.replace(key, value)

            # ---- LPC-3. Transliteration ----
            self._transliterated_tokens[section_num] = [epi.transliterate(token).replace('6', '').replace('4', '')
                                                        for token in self._stressed_tokens[section_num]]
            for i, token in enumerate(self._transliterated_tokens[section_num]):
                for key, value in non_ipa_symbols.items():
                    if key in token:
                        token = token.replace(key, value)
                        self._transliterated_tokens[section_num][i] = token

            # ---- LPC-4. Common rules ----
            # fricative g
            for i, token in enumerate(self._transliterated_tokens[section_num]):
                try:
                    next_token = self._transliterated_tokens[section_num][i + 1]
                except IndexError:
                    next_token = ' '

                token_let = self._tokens[section_num][i]
                nlp_token = nlp(token_let)[0]
                lemma = nlp_token.lemma_

                if lemma in {'ага', 'ого', 'угу', 'господь', 'господи', 'бог'}:
                    self._transliterated_tokens[section_num][i] = token.replace('ɡ', 'γ', 1)
                elif token_let in {'ах', 'эх', 'ох', 'ух'}:
                    next_token_allophone = allophones.get(next_token[0], {})
                    if next_token_allophone.get('voice', '') == 'voiced':
                        self._transliterated_tokens[section_num][i] = token.replace('x', 'γ', 1)

            # ---- Join phonemes ----
            section_phonemes_list = []
            joined_tokens = '_'.join(self._transliterated_tokens[section_num])
            i = 0
            counter = 0
            default_len = len(joined_tokens)
            while i < default_len:
                if joined_tokens[i] not in ['+', '-']:
                    n = 4
                    if i != default_len - 1:
                        while (joined_tokens[i: i + n] not in epi_starterpack + ['_', '|', '||', 'γ', 'ʐ']) and (n > 0):
                            counter += 1
                            if counter > 10000:
                                raise IndexError('Endless loop')
                            n -= 1
                        section_phonemes_list.append(joined_tokens[i: i + n])
                    elif joined_tokens[i] in epi_starterpack + ['||', 'γ']:
                        section_phonemes_list.append(joined_tokens[i])
                    i += n
                else:
                    section_phonemes_list.append(joined_tokens[i])
                    i += 1

            section_phonemes_list = [x for x in section_phonemes_list if x not in ['', 'ʲ']]
            self._phonemes.append(section_phonemes_list)

            n = 0
            for allophone_index in range(len(self._phonemes[section_num]) - 1):
                allophone = self._phonemes[section_num][allophone_index + n]
                next_allophone = self._phonemes[section_num][allophone_index + n + 1]
                if (allophone == 't͡s' and next_allophone == 's') or (allophone == 'd͡ʒ' and next_allophone == 'ʐ'):
                    del self._phonemes[section_num][allophone_index + n + 1]
                    n -= 1

            # ---- Join letters ----
            self._letters_list.append(list('_'.join(self._stressed_tokens[section_num])))

            # ---- Continue LPC-4. Common rules ----
            self._phonemes[section_num] = fix_jotised(self._phonemes[section_num],
                                                      self._letters_list[section_num])
            self._phonemes[section_num] = shch(self._phonemes[section_num])
            self._phonemes[section_num] = long_ge(self._phonemes[section_num])
            self._phonemes[section_num] = assimilative_palatalization(self._tokens[section_num],
                                                                      self._phonemes[section_num])
            self._phonemes[section_num] = long_consonants(self._phonemes[section_num])
            self._phonemes[section_num] = stunning(self._phonemes[section_num])

            # ---- Allophones ----
            self._allophones[section_num] = nasal_m_n(self._phonemes[section_num])
            self._allophones[section_num] = silent_r(self._allophones[section_num])
            self._allophones[section_num] = voiced_ts(self._allophones[section_num])

            # ---- Extract phrasal words ----
            self._phrasal_words[section_num] = extract_phrasal_words(self._allophones[section_num],
                                                                     self._phrasal_words_indexes[section_num])
            # ---- Prestressed syllable ----
            section = self._phrasal_words[section_num][:]
            n = 0
            for symb_i, symb in enumerate(section):
                if symb == '+':
                    preavi = [phon_i for phon_i, phon in enumerate(section[:symb_i - 1]) if
                              allophones[phon]['phon'] == 'V' and '_' not in section[phon_i + n:symb_i]]
                    if preavi:
                        self._phrasal_words[section_num].insert(preavi[-1] + n + 1, '-')
                        n += 1

            #  ---- Vowels ----
            self._allophones[section_num] = vowels(self._phrasal_words[section_num])

            # ---- Labialization and velarization ----
            self._allophones[section_num] = labia_velar(self._allophones[section_num])

    def _insert_pauses(self, sounds_list: list):
        for i, key in enumerate(self._pause_dict):
            sounds_list.insert(i + key, self._pause_dict[key])

    def _get_escape_symbols(self, save_stresses: bool = False, save_spaces: bool = False):
        escape_symbols = ['+', '-', '_']
        if save_stresses:
            escape_symbols.remove('+')
        if save_spaces:
            escape_symbols.remove('_')

        return escape_symbols

    def _join_sounds(self, escape_symbols: list, sounds_list: list):
        return ' '.join(
            [' '.join([x for x in section if x not in escape_symbols])
             if section != '||'
             else section
             for section in sounds_list]
        )

    def allophones(self, stress_place: str = None, save_stresses=False, save_spaces=False, save_pauses=False,
                   stress_symbol='+'):
        """
        :param stress_place: 'after' - to place the stress symbol after the stressed vowel,
            'before' - to place the stress symbol before the stressed vowel.
        :param stress_symbol: A symbol that you want to indicate the stress.
            Be careful not to use letters and signs from the following list
            ['.', '_', '-', 'ʲ', 'ᶣ', 'ʷ', 'ˠ', 'ː', '͡']!
        :param save_spaces: Will replace spaces with '_'.
        :param save_stresses: Will replace stresses with the stress_symbol.
        :param save_pauses: Will replace punctuation with '||' for long pauses ('.', '?', '!', '…')
            and '|' for short pauses (other symbols).
        :return: List of allophones.
        """
        if save_pauses:
            self._insert_pauses(self._allophones)

        escape_symbols = self._get_escape_symbols(save_stresses=save_stresses, save_spaces=save_spaces)
        res = self._join_sounds(escape_symbols, self._allophones).split()

        if stress_place is None:
            stress_place = self._stress_place
        if stress_place == 'before':
            res = replace_stress_before(res)

        if (stress_symbol != '+') and ('+' not in escape_symbols):
            res = [x.replace('+', stress_symbol) for x in res]

        return res

    def phonemes(self, stress_place: str = None, save_stresses: bool = False, save_spaces: bool = False,
                 save_pauses: bool = False, stress_symbol: str = '+'):
        """
        :param stress_place: 'after' - to place the stress symbol after the stressed vowel,
            'before' - to place the stress symbol before the stressed vowel.
        :param stress_symbol: A symbol that you want to indicate the stress.
            Be careful not to use letters and signs from the following list
            ['.', '_', '-', 'ʲ', 'ᶣ', 'ʷ', 'ˠ', 'ː', '͡']!
        :param save_spaces: Will replace spaces with '_'.
        :param save_stresses: Will replace stresses with the stress_symbol.
        :param save_pauses: Will replace punctuation with '||' for long pauses ('.', '?', '!', '…')
            and '|' for short pauses (other symbols).
        :return: List of phonemes.
        """
        if save_pauses:
            self._insert_pauses(self._phonemes)

        escape_symbols = self._get_escape_symbols(save_stresses=save_stresses, save_spaces=save_spaces)
        res = self._join_sounds(escape_symbols, self._phonemes).split()

        if stress_place is None:
            stress_place = self._stress_place
        if stress_place == 'before':
            res = replace_stress_before(res)

        if (stress_symbol != '+') and ('+' not in escape_symbols):
            res = [x.replace('+', stress_symbol) for x in res]

        return res

    def stressed_text(self, stress_place: str = None, stress_symbol: str = '+'):
        """
        :param stress_place: 'after' - to place the stress symbol after the stressed vowel,
            'before' - to place the stress symbol before the stressed vowel.
        :param stress_symbol: A symbol that you want to indicate the stress.
            Be careful not to use letters and signs from the following list
            ['.', '_', '-', 'ʲ', 'ᶣ', 'ʷ', 'ˠ', 'ː', '͡']!
        :return: A text string with stresses.
        """
        if stress_place is None:
            stress_place = self._stress_place
        if stress_place == 'before':
            res = ' '.join([' '.join(replace_stress_before(section)) for section in self._stressed_text])
        else:
            res = ' '.join([' '.join(section) for section in self._stressed_text])

        if stress_symbol != '+':
            res = [x.replace('+', stress_symbol) for x in res]

        return res
